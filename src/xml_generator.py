import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np
import os
from datetime import datetime,timedelta
from bs4 import BeautifulSoup
import lxml.etree as etree
import xml.dom.minidom
from xml.dom.minidom import parse
from openpyxl import load_workbook
import json
from datetime import date
import xlwings as xw


col =['tipoCartera', 'programaCredito', 'tipoOperacion', 'tipoMoneda', 'tipoAgrupamiento', 'numeroPagare',
        'numeroObligacionIntermediario', 'fechaSuscripcion', 'fechaDesembolso', 'oficinaPagare', 'oficinaObligacion',
        'codigo', 'cantidad', 'correoElectronico','cumpleCondicionesProductorAgrupacion', 'tipoAgrupacion', 'tipoPersona',
        'tipoProductor', 'actividadEconomica', 'tipo', 'numeroIdentificacion', 'digitoVerificacion', 'primerApellido',
        'SegundoApellido', 'PrimerNombre', 'SegundoNombre', 'Razonsocial', 'direccion', 'municipio', 'prefijo',
        'numero', 'valor', 'fechaCorte', 'fechaInicialEjecucion', 'fechaFinalEjecucion', 'tipo2', 'municipio2', 'direccion2',
         'codigo2', 'unidadesAFinanciar', 'costoInversion', 'valorAFinanciar', 'fechaVencimientoFinal', 'plazoCredito',
         'valorTotalCredito', 'porcentaje', 'valorObligacion', 'registro', 'fechaAplicacionHasta', 'conceptoRegistroCuota',
          'periodicidadIntereses', 'periodicidadCapital', 'tasaBaseBeneficiario', 'margenTasaBeneficiario', 'valorCuotaCapital',
          'porcentajeCapitalizacionIntereses', 'margenTasaRedescuento','valorIngresos','Anticipo','valorTotalProyecto','valorTotalAFinanciar',
          'plazoFinanciacion','numeroProyecto','numeroDesembolso','desembolsos']


archivo = os.path.join('data', 'Prueba2.xlsx')
tabla=pd.read_excel(archivo,names=col,index_col=False)
#print(tabla)
#registros=pd.read_excel(archivo,names=col,index_col=False,sheet_name='Hoja3')


tabla.fillna("", inplace=True)
#abla["tipoCartera"] = tabla["tipoCartera"].astype(float)
tabla['fechaSuscripcion']=tabla['fechaSuscripcion'].astype(str)
tabla['fechaSuscripcion'] = tabla['fechaSuscripcion'].str.replace('/', '-')
tabla['fechaDesembolso']=tabla['fechaDesembolso'].astype(str)
tabla['fechaDesembolso'] = tabla['fechaDesembolso'].str.replace('/', '-')
tabla['fechaCorte']= tabla['fechaCorte'].astype(str)
tabla['fechaCorte'] = tabla['fechaCorte'].str.replace('/', '-')
tabla['fechaInicialEjecucion']=tabla['fechaInicialEjecucion'].astype(str)
tabla['fechaInicialEjecucion'] = tabla['fechaInicialEjecucion'].str.replace('/', '-')
tabla['fechaFinalEjecucion']=tabla['fechaFinalEjecucion'].astype(str)
tabla['fechaFinalEjecucion'] = tabla['fechaFinalEjecucion'].str.replace('/', '-')
tabla['fechaAplicacionHasta']=tabla['fechaAplicacionHasta'].astype(str)
tabla['fechaAplicacionHasta'] = tabla['fechaAplicacionHasta'].str.replace('/', '-')
tabla['fechaVencimientoFinal']=tabla['fechaVencimientoFinal'].astype(str)
tabla['fechaVencimientoFinal'] = tabla['fechaVencimientoFinal'].str.replace('/', '-')

tabla['PrimerNombre'] = tabla['PrimerNombre'].str.replace('Ð', 'Ñ')
tabla['SegundoNombre'] = tabla['SegundoNombre'].str.replace('Ð', 'Ñ')
tabla['primerApellido'] = tabla['primerApellido'].str.replace('Ð', 'Ñ')
tabla['SegundoApellido'] = tabla['SegundoApellido'].str.replace('Ð', 'Ñ')
tabla['Razonsocial'] = tabla['Razonsocial'].str.replace('Ð', 'Ñ')
tabla['correoElectronico'] = tabla['correoElectronico'].str.replace('Ð', 'Ñ')
#tabla['numeroPagare'] = tabla['numeroPagare'].apply(lambda x: x.replace('F','00'))
#tabla['numeroObligacionIntermediario'] = tabla['numeroObligacionIntermediario'].apply(lambda x: x.replace('F','00'))
tabla['digitoVerificacion']=tabla['digitoVerificacion'].astype(str)
tabla['digitoVerificacion'] = tabla['digitoVerificacion'].apply(lambda x: x.replace('.0',''))
tabla['numero']=tabla['numero'].astype(str)
tabla['numero'] = tabla['numero'].apply(lambda x: x.replace('.0',''))
tabla['municipio']=tabla['municipio'].astype(str)
tabla['municipio'] = tabla['municipio'].apply(lambda x: x.replace('.0',''))



nur =  input('Cuantas operaciones vas a cargar: ')
#vtotal =  input('Cual es el valor total de la carga: ')
vtotal = tabla['valorTotalCredito'].sum()
vtotal = vtotal/2
nur=str(nur)
vtotal=str(vtotal)
vtotal = vtotal.replace('.0','')

#ET.register_namespace('xsi', "http://www.w3.org/2001/XMLSchema-instance")
#ET.register_namespace('xsd', "http://www.w3.org/2001/XMLSchema")


ET.register_namespace('', "http://www.finagro.com.co/sit")
Obligaciones = ET.Element('obligaciones', {'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance',
                                           'xmlns:xsd': 'http://www.w3.org/2001/XMLSchema',
                                           'cifraDeControl': nur,
                                           'cifraDeControlValor': vtotal,
                                           
                                           })

#Obligaciones=ET.Element('{http://www.finagro.com.co/sit}obligaciones',cifraDeControlValor=vtotal,cifraDeControl=nur)
#Obligaciones=ET.Element('{http://www.w3.org/2001/XMLSchema}obligaciones',cifraDeControlValor=vtotal,cifraDeControl=nur)
Ni = len(tabla['tipoCartera'])
#print(tabla)

for i in range(Ni):

    tc=tabla.iloc[i,0]  #tipoCartera
    pc=tabla.iloc[i,1]  #programaCredito
    to=tabla.iloc[i,2]  #tipoOperacion
    tm=tabla.iloc[i,3]  #tipoMoneda
    ta=tabla.iloc[i,4]  #tipoAgrupamiento
    np=tabla.iloc[i,5]  #numeroPagare
    noi=tabla.iloc[i,6] #numeroObligacionIntermediario
    fs=tabla.iloc[i,7]  #fechaSuscripcion
    fd=tabla.iloc[i,8]  #fechaDesembolso
    op=tabla.iloc[i,9]  #oficinaPagare
    ofo=tabla.iloc[i,10]    #oficinaObligacion
    cod=tabla.iloc[i,11]    #codigo
    can=tabla.iloc[i,12]    #cantidad
    email=tabla.iloc[i,13]  #correoElectronico
    cpa=tabla.iloc[i,14]
    ta2=tabla.iloc[i,15]    #tipoAgrupacion
    tp=tabla.iloc[i,16]     #tipoPersona
    tpr=tabla.iloc[i,17]    #tipoProductor
    ae=tabla.iloc[i,18] #actividadEconomica
    tipo=tabla.iloc[i,19]   #tipo
    id=tabla.iloc[i,20] #numeroIdentificacion
    dv=tabla.iloc[i,21] #digitoVerificacion
    pa=tabla.iloc[i,22] #PrimerApellido
    sa=tabla.iloc[i,23] #SegundoApellido
    pn=tabla.iloc[i,24] #PrimerNombre
    sn=tabla.iloc[i,25] #SegundoNombre
    rs=tabla.iloc[i,26] #Razonsocial
    dir=tabla.iloc[i,27]    #direccion
    mun=tabla.iloc[i,28]    #municipio
    pref=tabla.iloc[i,29]   #prefijo
    num=tabla.iloc[i,30]    #numero
    va=tabla.iloc[i,31] #valor
    fc=tabla.iloc[i,32] #fechaCorte
    fie=tabla.iloc[i,33]    #fechaInicialEjecucion
    ffe=tabla.iloc[i,34]    #fechaFinalEjecucion
    tipo2=tabla.iloc[i,35]  #tipo2
    mun2=tabla.iloc[i,36]   #municipio3
    dir2=tabla.iloc[i,37]   #direccion4
    cod2=tabla.iloc[i,38]   #codigo5
    uf=tabla.iloc[i,39] #unidadesAFinanciar
    ci=tabla.iloc[i,40] #costoInversion
    vaf=tabla.iloc[i,41]    #valorAFinanciar
    fvf=tabla.iloc[i,42]    #fechaVencimientoFinal
    plazo=tabla.iloc[i,43]  #plazoCredito
    vtc=tabla.iloc[i,44]    #valorTotalCredito
    por=tabla.iloc[i,45]    #porcentaje
    vao=tabla.iloc[i,46]    #valorObligacion


    reg=tabla.iloc[i,47]    #registro
    fah=tabla.iloc[i,48]    #fechaAplicacionHasta
    crc=tabla.iloc[i,49]    #conceptoRegistroCuota
    pi=tabla.iloc[i,50] #periodicidadIntereses
    pec=tabla.iloc[i,51]    #periodicidadCapital
    tb=tabla.iloc[i,52] #tasaBaseBeneficiario
    mtb=tabla.iloc[i,53]    #margenTasaBeneficiario
    vcap=tabla.iloc[i,54]   #valorCuotaCapital
    pci=tabla.iloc[i,55]    #porcentajeCapitalizacionIntereses
    mtr=tabla.iloc[i,56]    #margenTasaRedescuento
    vai=tabla.iloc[i,57]
    #fci=tabla.iloc[i,58]
    pf=tabla.iloc[i,59] # La solicitud corresponde a  un proyecto financiado con varios desembolsos


    tc=str(tc)
    pc=str(pc)
    to=str(to)
    tm=str(tm)
    ta=str(ta)
    np=str(np)
    noi=str(noi)
    fs=str(fs)
    fd=str(fd)
    op=str(op)
    ofo=str(ofo)
    cod=str(cod)
    can=str(can)
    email=str(email)
    cpa=str(cpa)
    ta2=str(ta2)
    tp=str(tp)
    tpr=str(tpr)
    ae=str(ae)
    tipo=str(tipo)
    id=str(id)
    dv=str(dv)
    pn=str(pn)
    sn=str(sn)
    pa=str(pa)
    sa=str(sa)
    rs=str(rs)
    dir=str(dir)
    mun=str(mun)
    pref=str(pref)
    num=str(num)
    va=str(va)
    fc=str(fc)
    fie=str(fie)
    ffe=str(ffe)
    tipo2=str(tipo2)
    mun2=str(mun2)
    dir2=str(dir2)
    cod2=str(cod2)
    uf=str(uf)
    ci=str(ci)
    vaf=str(vaf)
    fvf=str(fvf)
    plazo=str(plazo)
    vtc=str(vtc)
    por=str(por)
    vao=str(vao)

    reg=str(reg)
    fah=str(fah)
    crc=str(crc)
    pi=str(pi)
    pec=str(pec)
    tb=str(tb)
    mtb=str(mtb)
    vcap=str(vcap)
    pci=str(pci)
    mtr=str(mtr)
    pf=str(pf)
    vai=str(vai)
    #fci=str(fci)


    if reg == str(1):

        
        obligacion=ET.SubElement(Obligaciones,'{http://www.finagro.com.co/sit}obligacion',tipoCartera=tc,programaCredito=pc,tipoOperacion=to,tipoMoneda=tm,tipoAgrupamiento=ta,numeroPagare=np,numeroObligacionIntermediario=noi,fechaSuscripcion=fs,fechaDesembolso=fd)
        ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}intermediario', oficinaPagare=op,oficinaObligacion=ofo,codigo=cod)
        beneficiarios=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}beneficiarios',cantidad=can)
        beneficiario=ET.SubElement(beneficiarios,'{http://www.finagro.com.co/sit}beneficiario', correoElectronico=email,cumpleCondicionesProductorAgrupacion="true",tipoAgrupacion=ta2,tipoPersona=tp,tipoProductor=tpr,actividadEconomica=ae)
        type(int(can))
        if dv == "":
            ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}identificacion', tipo=tipo, numeroIdentificacion=id)
            ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}nombre', primerNombre=pn, segundoNombre=sn, primerApellido=pa, segundoApellido=sa)

        else:
            ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}identificacion', tipo=tipo, numeroIdentificacion=id, digitoVerificacion=dv)
            ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}nombre', primerNombre=pn, segundoNombre=sn, primerApellido=pa, segundoApellido=sa,Razonsocial=rs)

        ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}direccionCorrespondencia', direccion=dir, municipio=mun)
        ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}numeroTelefono', prefijo=pref, numero=num)
        ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}valorActivos', valor=va, fechaCorte=fc, tipoDato="COP")
        ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}valorIngresos', valor=vai, fechaCorte='2023-12-31', tipoDato="COP")

        if pf == str(1): #(nuevo)
            proyecto=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}proyecto', fechaInicialEjecucion=fie, fechaFinalEjecucion=ffe)
            ET.SubElement(proyecto,'{http://www.finagro.com.co/sit}incentivo',inscripcionIncentivo="false")
            proyectosFinanciados=ET.SubElement(proyecto,'{http://www.finagro.com.co/sit}proyectosFinanciados',valorTotalProyecto="",valorTotalAFinanciar="",plazoFinanciacion="",numeroProyecto="",numeroDesembolso="99",desembolsos="")
            ET.SubElement(proyectosFinanciados,'{http://www.finagro.com.co/sit}destinosProyecto',codigoDestinoCredito="",valorAFinanciar="",unidadesAFinanciar="",costoInversion="")
            ET.SubElement(proyectosFinanciados,'{http://www.finagro.com.co/sit}municipios',municipio="")



        elif len(pf) == 0:
            proyecto=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}proyecto', fechaInicialEjecucion=fie, fechaFinalEjecucion=ffe)

        predios=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}predios')
        ET.SubElement(predios,'{http://www.finagro.com.co/sit}predio', tipo=tipo2, municipio=mun, direccion=dir)
        destinosCredito=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}destinosCredito')
        destinoCredito=ET.SubElement(destinosCredito,'{http://www.finagro.com.co/sit}destinoCredito', codigo=cod2, unidadesAFinanciar=uf, costoInversion=ci)
        destinoCreditoValorAFinanciar=ET.SubElement(destinoCredito,'{http://www.finagro.com.co/sit}destinoCreditoValorAFinanciar')
        ET.SubElement(destinoCreditoValorAFinanciar,'{http://www.finagro.com.co/sit}valorAFinanciar',xmlns="").text=vaf
        ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}financiacion',fechaVencimientoFinal=fvf,plazoCredito=plazo,valorTotalCredito=vtc,porcentaje=por,valorObligacion=vao)
        planPagos=ET.SubElement(obligacion,'{http://www.finagro.com.co/sit}planPagos')
        ET.SubElement(planPagos,'{http://www.finagro.com.co/sit}registroCuota',registro=reg,fechaAplicacionHasta=fah,conceptoRegistroCuota=crc,periodicidadIntereses=pi,periodicidadCapital=pec,tasaBaseBeneficiario=tb,margenTasaBeneficiario=mtb,valorCuotaCapital=vcap,porcentajeCapitalizacionIntereses=pci,margenTasaRedescuento=mtr)

    elif reg == str(2) or reg == str(3):
        ET.SubElement(planPagos,'{http://www.finagro.com.co/sit}registroCuota',registro=reg,fechaAplicacionHasta=fah,conceptoRegistroCuota=crc,periodicidadIntereses=pi,periodicidadCapital=pec,tasaBaseBeneficiario=tb,margenTasaBeneficiario=mtb,valorCuotaCapital=vcap,porcentajeCapitalizacionIntereses=pci,margenTasaRedescuento=mtr)

    # elif >= 2:
    #     type(str(can))
    #     for j in range(len(can)):
    #         beneficiario=ET.SubElement(beneficiarios,'{http://www.finagro.com.co/sit}beneficiario', correoElectronico=email,cumpleCondicionesProductorAgrupacion="true", tipoAgrupacion=ta2, tipoPersona=tp, tipoProductor=tpr, actividadEconomica=ae)
    #         ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}identificacion', tipo=tipo, numeroIdentificacion=id, digitoVerificacion=dv)
    #         ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}nombre', primerNombre=pn, segundoNombre=sn, primerApellido=pa, segundoApellido=sa, Razonsocial=rs)
    #         ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}direccionCorrespondencia', direccion=dir, municipio=mun)
    #         ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}numeroTelefono', prefijo=pref, numero=num)
    #         ET.SubElement(beneficiario,'{http://www.finagro.com.co/sit}valorActivos', valor=va, fechaCorte=fc)
    #

# xml = ET.tostring(Obligaciones)  # binary string
# #my_arr=xml.decode()
#
# with open('output2.xml', 'w') as f:  # Write in XML file as utf-8
#     f.write('<?xml version="1.0" encoding="UTF-8"?>' + xml.decode('utf-8'))

xml_str = ET.tostring(Obligaciones, encoding='unicode')
output = os.path.join('data', 'output2.xml')
with open(output, 'w', encoding='utf-8') as f:
    f.write('<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str)



# Formatear el archivo XML con codificación UTF-8
#file_path = r'C:\Users\arredondoivan\Downloads\Desarrollo Obligaciones Nuevas'
def format_xml(file_path):
    # Cargar el archivo XML
    with open(file_path, 'r', encoding='utf-8') as file:
        xml_str = file.read()

    # Usar minidom para parsear el contenido del archivo
    xml_doc = xml.dom.minidom.parseString(xml_str)

    # Formatear el contenido del archivo
    pretty_xml_str = xml_doc.toprettyxml()

    # Guardar el archivo formateado en el disco
    with open(file_path, 'w', encoding='utf-8') as file:

        file.write(pretty_xml_str)

# Formatear el archivo de salida
format_xml(output)
ciiu = pd.read_excel(os.path.join('data', 'Ejemplo.xlsx'))
ciiu = ciiu.astype('str')
data_dict = ciiu.to_dict(orient='records')

wb = load_workbook(filename=archivo)
hoja = wb['Hoja2']
archivo_2= r'M:\Bancos\Banco Caja Social\Informes\INFORME CAJA SOCIAL CONSOLIDADO.xlsx'
wb_informe = load_workbook(filename=archivo_2)
hoja_informe = wb_informe['REGISTRO PARA INFORME  DIARIO ']
archivo_3 = r'M:\Bancos\Banco Caja Social\Otros\PAF MICROCREDITO_CAJA SOCIAL.xlsx'
wb_planificacion = load_workbook(filename=archivo_3)
hoja_planificacion = wb_planificacion['FORMATO -PÁGINA 1 (1)']
hoja_planificacion_2 = wb_planificacion['Hoja1']

archivo_4 = os.path.join('data', 'Macro_organizacion_de_datos.xlsm')
wb_macro = load_workbook(filename=archivo_4)
hoja_macro = wb_macro['MICROCREDITO CAJA SOCIAL']


for row in range(1, hoja_informe.max_row + 1):
    if hoja_informe[f'A{row}'].value is not None:
        ultima_fila_con_texto = row
    
fecha = datetime.now()
fecha_formateada = fecha.strftime('%d-%m-%Y')
fecha_ayer = fecha - timedelta(days=1)
iterador = int((hoja.max_row)/2)
nur = int(nur)
for i in range(1,nur+1):
    wb_planificacion = load_workbook(filename=archivo_3)
    hoja_planificacion = wb_planificacion['FORMATO -PÁGINA 1 (1)']
    hoja_planificacion_2 = wb_planificacion['Hoja1']
    hoja_informe[f'A{ultima_fila_con_texto + i}'] = 'REGISTRADO'
    hoja_informe[f'B{ultima_fila_con_texto + i}'] = 'BANCO CAJA SOCIAL'
    hoja_informe[f'C{ultima_fila_con_texto + i}'] = hoja[f'U{i}'].value
    Nombre = '\PAFF CAJA SOCIAL_' + str(hoja[f'U{i}'].value) + '.xlsx'
    ruta = 'M:\Bancos\Banco Caja Social\Otros\Proyectos'+ Nombre
    hoja_planificacion['O9'] = hoja[f'U{i}'].value
    hoja_informe[f'AS{ultima_fila_con_texto + i}'] = hoja_macro[f'E{i + 1}'].value
    hoja_planificacion_2['A1'] = hoja_macro[f'E{i + 1}'].value
    cod_ciiu = str(hoja_macro[f'E{i + 1}'].value)

    descripcion = None
    for item in data_dict:
        if item['Codigo'] == cod_ciiu:
            descripcion = item['Descripción']
            break

    for j in range(3):  
        valores = []
        for letra in ['W', 'X', 'Y', 'Z']:  
            valor = hoja[f'{letra}{i}'].value
            if valor is None:
                valor = ''  
            valores.append(str(valor))

    justificacion = str('El objeto social de '+ ' '.join(valores)) + 'es ' + str(descripcion) 
    hoja_planificacion['C42'] = justificacion
    hoja_planificacion['C48'] = 'Los recursos de este proyecto serán utilizados en el crubrimiento de los costos y gastos relacionados con la ejecución del objeto social del cliente, principalmente ' + str(descripcion) 
    hoja_informe[f'D{ultima_fila_con_texto + i}'].value = ' '.join(valores)
    hoja_planificacion['C8'] = ' '.join(valores)
    hoja_informe[f'E{ultima_fila_con_texto + i}'] = hoja[f'AB{i}'].value
    hoja_planificacion['C10'] = hoja[f'AB{i}'].value
    #hoja_informe[f'F{ultima_fila_con_texto + i}'] = hoja[f'AC{i}'].value
    #hoja_planificacion['J10'] = hoja[f'AC{i}'].value
    hoja_informe[f'F{ultima_fila_con_texto + i}'] = str(hoja_macro[f'G{i + 1}'].value).upper()
    hoja_informe[f'G{ultima_fila_con_texto + i}'] = str(hoja_macro[f'I{i + 1}'].value).upper()
    hoja_planificacion['J10'] = str(hoja_macro[f'G{i + 1}'].value).upper()
    hoja_planificacion['K10'] = str(hoja_macro[f'I{i + 1}'].value).upper()
    hoja_informe[f'H{ultima_fila_con_texto + i}'] = hoja[f'AG{i}'].value
    hoja_informe[f'I{ultima_fila_con_texto + i}'] = hoja[f'AF{i}'].value
    hoja_planificacion['M14'] = hoja[f'AF{i}'].value 
    hoja_informe[f'J{ultima_fila_con_texto + i}'] = fecha_formateada
    hoja_informe[f'K{ultima_fila_con_texto + i}'] = fecha_ayer
    hoja_informe[f'L{ultima_fila_con_texto + i}'] = hoja[f'AP{i}'].value
    hoja_planificacion['J61'] = hoja[f'AP{i}'].value
    hoja_planificacion['K61'] = hoja[f'AP{i}'].value
    hoja_informe[f'M{ultima_fila_con_texto + i}'] = hoja[f'AP{i}'].value
    hoja_informe[f'N{ultima_fila_con_texto + i}'] = hoja[f'AR{i}'].value
    hoja_planificacion['M61'] = hoja[f'AR{i}'].value
    hoja_informe[f'O{ultima_fila_con_texto + i}'] = 'MES VENCIDO'
    hoja_informe[f'P{ultima_fila_con_texto + i}'] = 'MES VENCIDO'
    hoja_informe[f'Q{ultima_fila_con_texto + i}'] = hoja[f'BB{i}'].value
    hoja_planificacion['R61'] = hoja[f'BB{i}'].value
    if hoja[f'R{i}'].value == '11':
        hoja_informe[f'R{ultima_fila_con_texto + i}'] = 'MICROEMPRESARIO'
    else:
        hoja_informe[f'R{ultima_fila_con_texto + i}'] = 'MICROEMPRESARIO PPIB'

    hoja_informe[f'AQ{ultima_fila_con_texto + i}'] = hoja[f'F{i}'].value
    hoja_informe[f'AV{ultima_fila_con_texto + i}'] = hoja[f'BF{i}'].value
    hoja_planificacion['H14'] = hoja[f'BF{i}'].value
    hoja_planificacion['N2'] = hoja[f'G{i}'].value

    wb = xw.Book(archivo_4)
    hoja_macro = wb.sheets['MICROCREDITO CAJA SOCIAL']
    valor_ingreso = hoja_macro.range(f'AT{i+1}').value
    wb_macro = load_workbook(filename=archivo_4)
    hoja_macro = wb_macro['MICROCREDITO CAJA SOCIAL']
    hoja_planificacion['H13'] = 'Monto de Ingresos ' + valor_ingreso
    wb = xw.Book(archivo_4)
    hoja_macro = wb.sheets['MICROCREDITO CAJA SOCIAL']
    valor_ingreso = hoja_macro.range(f'AQ{i+1}').value
    wb_macro = load_workbook(filename=archivo_4)
    hoja_macro = wb_macro['MICROCREDITO CAJA SOCIAL']
    hoja_planificacion['M13'] = 'Monto de Activos ' + valor_ingreso
    
    wb_planificacion.save(filename=ruta)


wb_informe.save(filename=archivo_2)
wb_macro.close()
print('PROCESO FINALIZADO SATISFACTORIAMENTE !!!')
